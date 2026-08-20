"""Import manually recorded token offers from the first two sheets of materials.xlsx."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
CYRILLIC_TRANSLIT = str.maketrans({
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "yo", "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
})
COLORS = {"kras": "red", "zel": "green", "sin": "blue"}
FAMILIES = {"stat": "stat", "kach": "quality", "svo": "trait"}
SCOPES = {"perv": "first", "posl": "last", "sluch": "random_one"}


def _normalize(value: object) -> str:
    text = str(value or "").strip().lower().translate(CYRILLIC_TRANSLIT)
    text = " ".join(text.split())
    return text.replace("sduch", "sluch").replace("sov", "svo").replace("kars", "kras")


def token_id(raw: object) -> str:
    text = _normalize(raw)
    compact = text.replace(" ", "")
    if compact in {"sluch1", "sluch+1"}:
        return "quality_shift_plus1"
    if compact in {"sluch2-1", "sluch2minus1"}:
        return "quality_shift_plus2_minus1"
    parts = text.split()
    if len(parts) < 2 or parts[0] not in FAMILIES or parts[1] not in COLORS:
        raise ValueError(f"Unrecognized materials token label: {raw!r} -> {text!r}")
    family, color = FAMILIES[parts[0]], COLORS[parts[1]]
    scope = "all" if len(parts) == 2 else SCOPES.get(parts[2])
    if scope is None:
        raise ValueError(f"Unrecognized materials token scope: {raw!r} -> {text!r}")
    return f"{family}_{scope}_{color}"


def main() -> None:
    parser = argparse.ArgumentParser(description="Import sheets 1 and 2 from a manual materials workbook.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--csv-output", default=str(ROOT / "data" / "raw" / "rng_token_observations" / "materials_sheets_1_2.csv"))
    parser.add_argument("--report-output", default=str(ROOT / "reports" / "rng_token_materials_sheets_1_2_comparison.json"))
    parser.add_argument("--empirical-preset", default=str(ROOT / "configs" / "rng_tokens" / "observed_run1_8_empirical_v1.json"))
    args = parser.parse_args()

    workbook = load_workbook(args.input, data_only=True)
    if len(workbook.worksheets) < 3:
        raise RuntimeError("Expected a metadata sheet plus two token-record sheets.")
    records: list[dict[str, object]] = []
    for sheet_index, sheet in enumerate(workbook.worksheets[1:3], start=1):
        for row in sheet.iter_rows(values_only=True):
            cells = list(row)
            roll = next((int(value) for value in cells[3:] if isinstance(value, (int, float))), None)
            if roll is None:
                continue
            for offer_position, raw in enumerate(cells[:3], start=1):
                if raw is None or not str(raw).strip():
                    continue
                records.append({
                    "source_file": Path(args.input).name,
                    "sheet_index": sheet_index,
                    "sheet_name": sheet.title,
                    "roll_number": roll,
                    "offer_position": offer_position,
                    "raw_label": str(raw),
                    "token_id": token_id(raw),
                })
    if len(records) != 180:
        raise RuntimeError(f"Expected 180 offers from two 30-roll sheets, received {len(records)}.")

    csv_path = Path(args.csv_output); csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(records[0]))
        writer.writeheader(); writer.writerows(records)

    observed = Counter(str(record["token_id"]) for record in records)
    preset = json.loads(Path(args.empirical_preset).read_text(encoding="utf-8"))
    existing = {str(row["token_id"]): int(row.get("exact_observed_count", 0)) for row in preset["token_offer_distribution"]}
    existing_total = sum(existing.values())
    merged_total = existing_total + len(records)
    rows = []
    for key in sorted(set(existing) | set(observed)):
        prior_count = existing.get(key, 0); added_count = observed.get(key, 0); merged_count = prior_count + added_count
        rows.append({
            "token_id": key,
            "current_empirical_count": prior_count,
            "current_empirical_probability": prior_count / existing_total if existing_total else 0.0,
            "materials_count": added_count,
            "materials_probability": added_count / len(records),
            "merged_count": merged_count,
            "merged_probability": merged_count / merged_total,
            "count_delta": added_count,
            "probability_delta": (merged_count / merged_total) - (prior_count / existing_total) if existing_total else 0.0,
        })
    report = {
        "source": {"input": str(Path(args.input).resolve()), "sheets": [workbook.worksheets[1].title, workbook.worksheets[2].title]},
        "records_saved": len(records),
        "records_csv": str(csv_path.resolve()),
        "current_empirical_total": existing_total,
        "materials_total": len(records),
        "merged_total": merged_total,
        "rows": rows,
        "quality_shift_summary": {
            "current_count": sum(existing.get(key, 0) for key in ("quality_shift_plus1", "quality_shift_plus2_minus1")),
            "materials_count": sum(observed.get(key, 0) for key in ("quality_shift_plus1", "quality_shift_plus2_minus1")),
            "merged_count": sum(existing.get(key, 0) + observed.get(key, 0) for key in ("quality_shift_plus1", "quality_shift_plus2_minus1")),
        },
        "note": "Raw spreadsheet observations are preserved separately. This import does not overwrite the immutable empirical preset.",
    }
    report_path = Path(args.report_output); report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({key: report[key] for key in ("records_saved", "current_empirical_total", "materials_total", "merged_total", "quality_shift_summary")}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
